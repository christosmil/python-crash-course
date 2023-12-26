from openpyxl import load_workbook


def get_all_rows(data_file, worksheet_name):
    """Get all rows from the given workbook and worksheet."""
    # Load the entire workbook.
    wb = load_workbook(data_file, data_only=True)

    # List all the sheets in the file.
    print("Found the following worksheets:")
    for sheetname in wb.sheetnames:
        print(sheetname)

    # Load one worksheet.
    ws = wb[worksheet_name]
    all_rows = list(ws.rows)

    return all_rows

def summarize_data(all_rows):
    """Summarize demographic data for police killings of African Americans,
    for each state in the dataset.
    """
    print(f"\nFound {len(all_rows)} rows of data.")

    print("\nFirst rows of data:")
    for row in all_rows[:5]:
        print(row)

    for cell in all_rows[0]:
        print(cell.value)

    for row in all_rows[1:52]:
        state = row[0].value
        percent_aa = int(round(row[3].value, 2) * 100)
        percent_aa_victims = int(round(row[4].value, 2) * 100)

        print(f"\n{state}")
        print(f" {percent_aa}% of residents are African American")
        print(f" {percent_aa_victims}% killed by police were African American")


data_file = 'data/mapping_police_violence_snapshot_061920.xlsx'
data = get_all_rows(data_file, '2013-2019 Killings by State')
summarize_data(data)